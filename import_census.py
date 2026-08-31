#!/usr/bin/env python3
"""
Census import script — Nottingham Park Houses
Imports 1911 and 1921 census data from Excel files into the PostgreSQL database.
Existing records are never overwritten (skipped on conflict).

Usage:
  pip install openpyxl psycopg2-binary --break-system-packages
  DATABASE_URL=postgres://... python3 import_census.py

Place the two Excel files alongside this script (or pass paths as arguments):
  python3 import_census.py 1911censusmasterlinzij.xlsx 1921censuspark.xlsx
"""

import os, sys, re, json
import openpyxl

try:
    import psycopg2
except ImportError:
    print("Installing psycopg2-binary...")
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'psycopg2-binary', '--break-system-packages', '-q'])
    import psycopg2

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

FILE_1911 = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, '1911censusmasterlinzij.xlsx')
FILE_1921 = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, '1921censuspark.xlsx')
ALL_PROPS_FILE = os.path.join(SCRIPT_DIR, 'data', 'all_props.json')

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    print("ERROR: DATABASE_URL environment variable is required.")
    print("  Get it from Railway → your project → Variables tab.")
    print("  Run: DATABASE_URL='postgres://...' python3 import_census.py")
    sys.exit(1)

# ── Property lookup ───────────────────────────────────────────────────────────

with open(ALL_PROPS_FILE) as f:
    ALL_PROPS = json.load(f)

def normalize(s):
    if s is None: return ''
    s = str(s).lower().strip()
    s = re.sub(r'\s+', ' ', s)
    return s

def normalize_street(s):
    s = normalize(s)
    # Expand common abbreviations / truncations
    subs = [
        (r'\bcav\b', 'cavendish'), (r'\bpk\b', 'park'),
        (r'\bcresn\b', 'crescent'), (r'\bcresc\b', 'crescent'),
        (r'\bcres\b', 'crescent'),
        (r'\brd\b', 'road'), (r'\bave\b', 'avenue'), (r'\bdr\b', 'drive'),
        (r'\bln\b', 'lane'), (r'\bterr\b', 'terrace'),
        (r'\bnort$', 'north'), (r'\bsout$', 'south'),
        (r'park house:\s*\d+\w*\s*', ''),  # strip "Park House: 14 " prefix
        (r'\bfish pond\b', 'fishpond'),    # "fish pond drive" → "fishpond drive"
        (r'\bs\.?\s*$', 'south'),          # trailing "S" or "S." = South (Cres S → Crescent South)
    ]
    for pat, rep in subs:
        s = re.sub(pat, rep, s)
    return s.strip()

# Build lookup: (normalized_no, normalized_street) → prop_id
# Also build street-only index for fuzzy matching
prop_by_no_street = {}
prop_by_street = {}   # normalized_street → list of (normalized_no, prop_id)

for p in ALL_PROPS:
    ns = normalize_street(p.get('street', ''))
    no = normalize(p.get('no', ''))
    pid = p['id']
    if ns and no:
        prop_by_no_street[(no, ns)] = pid
    if ns:
        prop_by_street.setdefault(ns, []).append((no, pid))

def find_property_id(raw_no, raw_street):
    """Return property_id or None. Tries exact then prefix-fuzzy match."""
    no = normalize(str(raw_no)) if raw_no else ''
    st = normalize_street(str(raw_street)) if raw_street else ''
    if not st: return None

    # 1. Exact
    if (no, st) in prop_by_no_street:
        return prop_by_no_street[(no, st)]

    # 2. Prefix fuzzy — census street may be truncated
    min_prefix = 8
    for prop_st, entries in prop_by_street.items():
        mlen = min(len(st), len(prop_st))
        if mlen < min_prefix: continue
        if st[:mlen] == prop_st[:mlen]:
            # Street matches — find entry with matching no
            for pno, pid in entries:
                if pno == no:
                    return pid
            # If only one property on that street, use it
            if len(entries) == 1:
                return entries[0][1]

    return None

# ── Relationship helpers ──────────────────────────────────────────────────────

# 1911 / 1921-B relationship flag columns relative to col offset
# 1911: cols 10-15 = HEAD, WIFE, daughter, son, Other, Servant
# 1921B: cols 13-18 = HEAD, WIFE, daughter, son, Other, Servant
REL_LABELS = ['Head', 'Wife', 'Daughter', 'Son', 'Other', 'Servant']

def rel_from_flags(row, start_col):
    """Derive relationship string from flag columns."""
    for i, label in enumerate(REL_LABELS):
        col = start_col + i
        if col < len(row) and row[col] and str(row[col]) not in ('None', '0', ''):
            val = str(row[col])
            if val == '1' or val.lower() == label.lower():
                return label
            if val not in ('0',):
                return val  # free-text relationship
    return None

def clean_str(v, falsy=('-', 'none', 'n/a', 'home duties', 'no occupation', '')):
    if v is None: return None
    s = str(v).strip()
    if s.lower() in falsy: return None
    return s

def clean_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

# ── DB helpers ────────────────────────────────────────────────────────────────

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

stats = dict(people_matched=0, people_created=0,
             census_inserted=0, census_skipped=0,
             no_address=0, rows_skipped=0)
address_misses = set()
unmatched_rows = []  # {year, name, no, street, age, relationship}

def find_or_create_person(first_name, last_name, born_year=None):
    fn = (first_name or '').strip()
    ln = (last_name or '').strip()
    if not fn or not ln: return None

    cur.execute(
        "SELECT id, born_year FROM people WHERE LOWER(first_name)=%s AND LOWER(last_name)=%s",
        (fn.lower(), ln.lower())
    )
    rows = cur.fetchall()
    if rows:
        if len(rows) == 1 or not born_year:
            stats['people_matched'] += 1
            return rows[0][0]
        # Multiple — pick closest birth year
        best = min(rows, key=lambda r: abs((r[1] or 0) - born_year) if r[1] else 999)
        stats['people_matched'] += 1
        return best[0]

    # Create
    cur.execute(
        "INSERT INTO people (first_name, last_name, born_year) VALUES (%s,%s,%s) RETURNING id",
        (fn, ln, born_year)
    )
    stats['people_created'] += 1
    return cur.fetchone()[0]

def insert_census(person_id, prop_id, year, relationship, age, occupation, source):
    cur.execute(
        "SELECT id FROM census_entries WHERE person_id=%s AND census_year=%s AND property_id=%s",
        (person_id, year, prop_id)
    )
    if cur.fetchone():
        stats['census_skipped'] += 1
        return
    cur.execute(
        """INSERT INTO census_entries
           (person_id, property_id, census_year, relationship, age_at_census, occupation_at_census, source)
           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
        (person_id, prop_id, year, relationship, age, occupation, source)
    )
    stats['census_inserted'] += 1

# ── 1911 import ──────────────────────────────────────────────────────────────
# Sheet layout (2 header rows then data):
# Col 0: House ID  1: House Name  2: NO.  3: Address(street)
# Col 4: Family(surname)  5: Name(first)  6-9: gender flags
# Col 10-15: rel flags (HEAD/WIFE/daughter/son/Other/Servant)
# Col 16: Age  Col 23: Occupation  Col 28: Born(Notts)  Col 29: Born(Other)

SKIP_1911 = {'Export Summary', 'template', 'comparison', 'Personaloccupation'}

def process_1911():
    print(f"\n=== 1911 Census ({FILE_1911}) ===")
    if not os.path.exists(FILE_1911):
        print(f"  File not found: {FILE_1911}"); return

    wb = openpyxl.load_workbook(FILE_1911, data_only=True)
    for sheet_name in wb.sheetnames:
        sl = sheet_name.lower()
        if sheet_name in SKIP_1911 or any(x in sl for x in ['template','compilation','comparison']):
            continue

        ws = wb[sheet_name]
        data_rows = [r for r in ws.iter_rows(values_only=True)]
        if len(data_rows) < 3: continue

        current_no = None
        current_street = None

        for row in data_rows[2:]:  # skip 2 header rows
            if all(v is None for v in row): continue

            house_no = row[2] if len(row) > 2 else None
            street   = row[3] if len(row) > 3 else None
            surname  = row[4] if len(row) > 4 else None
            forename = row[5] if len(row) > 5 else None
            age      = row[16] if len(row) > 16 else None
            occ      = row[23] if len(row) > 23 else None

            # Skip header-like or empty rows
            if not forename or not surname: stats['rows_skipped'] += 1; continue
            if str(forename).lower() in ('name','first name','first name(s)'): continue
            if str(surname).lower() in ('no entry','family'): continue
            # Skip dash-only or digit-only values (empty cells / totals rows)
            if str(surname).strip() in ('-',) or str(surname).strip().isdigit(): continue
            if str(forename).strip() in ('-',): continue

            # Carry forward house address ('-' is an empty Excel cell, not a value)
            SKIP_VALS = {'no entry','none','','-'}
            if house_no and str(house_no).lower() not in SKIP_VALS:
                current_no = str(house_no).strip()
            if street and str(street).lower() not in SKIP_VALS:
                current_street = str(street).strip()

            prop_id = find_property_id(current_no, current_street)
            if not prop_id:
                address_misses.add((str(current_no or ''), str(current_street or '')))
                relationship = rel_from_flags(row, 10)
                if len(row) > 14 and row[14] and str(row[14]) not in ('1','None','0',''):
                    relationship = str(row[14])
                unmatched_rows.append({'year': 1911,
                    'name': f"{forename} {surname}",
                    'no': current_no or '', 'street': current_street or '',
                    'age': clean_int(age), 'relationship': relationship})
                stats['no_address'] += 1; continue

            relationship = rel_from_flags(row, 10)
            # col 14 "Other" may hold free text
            if len(row) > 14 and row[14] and str(row[14]) not in ('1','None','0',''):
                relationship = str(row[14])

            age_val = clean_int(age)
            born_year = (1911 - age_val) if age_val else None
            occ_val = clean_str(occ)

            person_id = find_or_create_person(str(forename), str(surname), born_year)
            if not person_id: continue

            insert_census(person_id, prop_id, 1911, relationship,
                          age_val, occ_val, 'National Archives 1911 Census')

    conn.commit()
    print(f"  Committed. Running totals: {stats}")

# ── 1921 import ──────────────────────────────────────────────────────────────
# Format A (1921CCN, 1921pelham cres): 1 header row
#   Col 0: first name  1: last name  2: relationship  3: sex
#   Col 4: birth year  5: age  6: birthplace  7: occupation  8: employer
#   Col 9: house NO    10: street
#
# Format B (all other data sheets): 2 header rows
#   Col 0: House ID  1: year  2: count  3: Location(street)
#   Col 4: House Name  5: house NO  6: Family(surname)  7: Name(first)
#   Col 8: person count  9: gender  10-12: gender flags
#   Col 13-18: rel (HEAD/WIFE/daughter/son/Other/Servant)
#   Col 19: Age  Col 26: Occupation  Col 31-32: Born

SKIP_1921 = {'template', '1921ccntemplate', '1921north rdtemplate',
             '1921pelhamcrestemplate (2)'}
FORMAT_A_SHEETS = {'1921CCN', '1921pelham cres'}

def process_1921():
    print(f"\n=== 1921 Census ({FILE_1921}) ===")
    if not os.path.exists(FILE_1921):
        print(f"  File not found: {FILE_1921}"); return

    wb = openpyxl.load_workbook(FILE_1921, data_only=True)
    for sheet_name in wb.sheetnames:
        sl = sheet_name.lower()
        if sheet_name in SKIP_1921 or any(x in sl for x in ['template','compilation','comparison','analysis']):
            continue

        ws = wb[sheet_name]
        data_rows = [r for r in ws.iter_rows(values_only=True)]
        if len(data_rows) < 2: continue

        if sheet_name in FORMAT_A_SHEETS:
            _process_1921_format_a(data_rows)
        else:
            _process_1921_format_b(data_rows)

    conn.commit()
    print(f"  Committed. Running totals: {stats}")

def _process_1921_format_a(data_rows):
    """Format A: clean layout, 1 header row."""
    for row in data_rows[1:]:
        if all(v is None for v in row): continue
        if not row[0] or str(row[0]).lower() in ('first name(s)','first name'): continue

        forename   = row[0] if len(row) > 0 else None
        surname    = row[1] if len(row) > 1 else None
        rel        = row[2] if len(row) > 2 else None
        birth_year = row[4] if len(row) > 4 else None
        age        = row[5] if len(row) > 5 else None
        occ        = row[7] if len(row) > 7 else None
        house_no   = row[9] if len(row) > 9 else None
        street     = row[10] if len(row) > 10 else None

        if not forename or not surname: stats['rows_skipped'] += 1; continue

        prop_id = find_property_id(house_no, street)
        if not prop_id:
            address_misses.add((str(house_no or ''), str(street or '')))
            unmatched_rows.append({'year': 1921,
                'name': f"{forename} {surname}",
                'no': str(house_no or ''), 'street': str(street or ''),
                'age': clean_int(row[5] if len(row) > 5 else None),
                'relationship': clean_str(rel)})
            stats['no_address'] += 1; continue

        born_year = clean_int(birth_year)
        age_val   = clean_int(age)
        if not born_year and age_val: born_year = 1921 - age_val

        person_id = find_or_create_person(str(forename), str(surname), born_year)
        if not person_id: continue

        insert_census(person_id, prop_id, 1921,
                      clean_str(rel), age_val,
                      clean_str(occ), 'National Archives 1921 Census')

def _process_1921_format_b(data_rows):
    """Format B: same structure as 1911, 2 header rows."""
    current_no = None
    current_street = None

    for row in data_rows[2:]:
        if all(v is None for v in row): continue

        street   = row[3] if len(row) > 3 else None
        house_no = row[5] if len(row) > 5 else None
        surname  = row[6] if len(row) > 6 else None
        forename = row[7] if len(row) > 7 else None
        age      = row[19] if len(row) > 19 else None
        occ      = row[26] if len(row) > 26 else None

        # Skip header/summary rows and dash/digit-only cells
        if not forename or not surname: stats['rows_skipped'] += 1; continue
        if str(forename).lower() in ('name','first name(s)','largest household'): continue
        if str(surname).lower() in ('family','no entry'): continue
        if str(surname).strip() in ('-',) or str(surname).strip().isdigit(): continue
        if str(forename).strip() in ('-',): continue
        if str(row[0]).lower() == 'total': continue  # skip TOTAL summary rows

        # Carry forward address ('-' = empty Excel cell, not a value)
        SKIP_VALS = {'none', '', '-'}
        if street and str(street).lower() not in SKIP_VALS:
            current_street = str(street).strip()
        if house_no and str(house_no).lower() not in SKIP_VALS:
            current_no = str(house_no).strip()

        prop_id = find_property_id(current_no, current_street)
        if not prop_id:
            address_misses.add((str(current_no or ''), str(current_street or '')))
            # Capture relationship for the unmatched report
            _rel_raw = row[13] if len(row) > 13 else None
            _rel = str(_rel_raw).strip() if _rel_raw and str(_rel_raw) not in ('1','0','None','') else rel_from_flags(row, 13)
            unmatched_rows.append({'year': 1921,
                'name': f"{forename} {surname}",
                'no': current_no or '', 'street': current_street or '',
                'age': clean_int(age), 'relationship': _rel})
            stats['no_address'] += 1; continue

        # Relationship: col 13 often has text; cols 13-18 are flag cols
        rel_raw = row[13] if len(row) > 13 else None
        if rel_raw and str(rel_raw) not in ('1','0','None',''):
            relationship = str(rel_raw).strip()
        else:
            relationship = rel_from_flags(row, 13)

        # col 17 "Other" may have free text
        if len(row) > 17 and row[17] and str(row[17]) not in ('1','0','None',''):
            relationship = str(row[17])

        age_val   = clean_int(age)
        born_year = (1921 - age_val) if age_val else None
        occ_val   = clean_str(occ)

        person_id = find_or_create_person(str(forename), str(surname), born_year)
        if not person_id: continue

        insert_census(person_id, prop_id, 1921, relationship,
                      age_val, occ_val, 'National Archives 1921 Census')

# ── Run ───────────────────────────────────────────────────────────────────────

print("Connecting to database...")
try:
    process_1911()
    process_1921()

    print("\n" + "="*50)
    print("IMPORT COMPLETE")
    print("="*50)
    print(f"  People matched (existing):  {stats['people_matched']}")
    print(f"  People created (new):       {stats['people_created']}")
    print(f"  Census entries inserted:    {stats['census_inserted']}")
    print(f"  Census entries skipped:     {stats['census_skipped']}  (already existed)")
    print(f"  Rows skipped (no name):     {stats['rows_skipped']}")
    print(f"  Addresses not matched:      {stats['no_address']}")

    if address_misses:
        print(f"\nUnique addresses not matched to a property ({len(address_misses)}):")
        for no, st in sorted(address_misses, key=lambda x: (x[1] or '', x[0] or '')):
            print(f"  no={no!r:10}  street={st!r}")

    if unmatched_rows:
        import csv
        csv_path = os.path.join(SCRIPT_DIR, 'census_unmatched.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=['year','name','no','street','age','relationship'])
            w.writeheader()
            w.writerows(sorted(unmatched_rows, key=lambda r: (r['street'] or '', r['no'] or '', r['year'])))
        print(f"\nFull unmatched person list written to: census_unmatched.csv ({len(unmatched_rows)} rows)")
        print("Review this file to decide which addresses need adding to all_props.json")

except Exception as e:
    conn.rollback()
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()
finally:
    cur.close()
    conn.close()
