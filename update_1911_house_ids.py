#!/usr/bin/env python3
"""
Update census_house_id for 1911 unresolved census entries, splitting each
street into individual household groups (PD1/PD2, PeD1/PeD2, SR1/SR2, etc.)
"""

import os
import psycopg2
import openpyxl

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise SystemExit("Set DATABASE_URL environment variable")

# Read spreadsheet data
wb = openpyxl.load_workbook('1911censusmasterlinzij.xlsx', read_only=True, data_only=True)

# Build list of (house_id, last_name, first_name) from each sheet
entries = []

SHEETS = ['Park Drive', 'Peveril Drive', 'southrd p 168', 'Kenilworthrd p 259', 'North Rd']

for sheet in SHEETS:
    if sheet not in wb.sheetnames:
        print(f"WARNING: sheet '{sheet}' not found")
        continue
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3, values_only=True):
        hid = row[0]
        last = row[4]
        first = row[5]
        if not hid or str(hid) == 'TOTAL':
            continue
        if not first and not last:
            continue
        hid = str(hid).strip()
        last = str(last).strip() if last else ''
        first = str(first).strip() if first else ''
        if last and first:
            entries.append((hid, last, first))

print(f"Total spreadsheet entries: {len(entries)}")

conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

updated = 0
skipped = 0
not_found = []

for house_id, last_name, first_name in entries:
    # Try exact match first (case-insensitive), joining people table
    cur.execute("""
        SELECT ce.id, ce.census_house_id
        FROM census_entries ce
        JOIN people p ON p.id = ce.person_id
        WHERE ce.census_year = 1911
          AND ce.property_id IS NULL
          AND LOWER(TRIM(p.last_name)) = LOWER(%s)
          AND LOWER(TRIM(p.first_name)) = LOWER(%s)
    """, (last_name, first_name))
    rows = cur.fetchall()

    if not rows:
        # Try partial first name match (spreadsheet may have middle names)
        cur.execute("""
            SELECT ce.id, ce.census_house_id
            FROM census_entries ce
            JOIN people p ON p.id = ce.person_id
            WHERE ce.census_year = 1911
              AND ce.property_id IS NULL
              AND LOWER(TRIM(p.last_name)) = LOWER(%s)
              AND LOWER(TRIM(p.first_name)) LIKE LOWER(%s)
        """, (last_name, first_name.split()[0] + '%'))
        rows = cur.fetchall()

    if not rows:
        not_found.append(f"{house_id} | {last_name} | {first_name}")
        skipped += 1
        continue

    for row_id, current_hid in rows:
        if current_hid == house_id:
            # Already correct
            continue
        cur.execute("""
            UPDATE census_entries SET census_house_id = %s WHERE id = %s
        """, (house_id, row_id))
        updated += 1

conn.commit()
cur.close()
conn.close()

print(f"\nUpdated: {updated}")
print(f"Not found: {skipped}")
if not_found:
    print("\nNot found in DB:")
    for nf in not_found:
        print(f"  {nf}")
